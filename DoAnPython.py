import tkinter as tk
from tkinter import messagebox, ttk, filedialog
import mysql.connector
from tkcalendar import DateEntry
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import warnings

warnings.filterwarnings("ignore")


def connect_db():
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="Lam10042006@",
        database="qlgiaovien"
    )
def center_window(win, w=950, h=550):
    ws = win.winfo_screenwidth()
    hs = win.winfo_screenheight()
    x = (ws // 2) - (w // 2)
    y = (hs // 2) - (h // 2)
    win.geometry(f'{w}x{h}+{x}+{y}')

def load_data():
    for i in tree.get_children():
        tree.delete(i)
    conn = connect_db()
    try:
        cur = conn.cursor()
        cur.execute("SELECT * FROM giaovien")
        rows = cur.fetchall()
        for row in rows:
            tree.insert("", tk.END, values=row)
    except Exception as e:
        messagebox.showerror("Lỗi kết nối", str(e))
    finally:
        conn.close()

def clear_input():
    entry_ma.delete(0, tk.END)
    entry_holot.delete(0, tk.END)
    entry_ten.delete(0, tk.END)
    gender_var.set("Nam")
    date_entry.set_date("2000-01-01")
    cbb_mon.set("")
    entry_ma.config(state='normal')

def ThemGiaoVien():
    ma = entry_ma.get()
    holot = entry_holot.get()
    ten = entry_ten.get()
    gioitinh = gender_var.get()
    ngaysinh = date_entry.get_date()
    monday = cbb_mon.get()

    if ma == "" or holot == "" or ten == "":
        messagebox.showwarning("Thiếu dữ liệu", "Vui lòng nhập Mã, Họ và tên Giáo viên")
        return
    conn = connect_db()
    try:
        cur = conn.cursor()
        cur.execute("INSERT INTO giaovien VALUES (%s, %s, %s, %s, %s, %s)",
        (ma, holot, ten, gioitinh, ngaysinh, monday))
        conn.commit()
        messagebox.showinfo("Thành công", "Thêm giáo viên thành công")
        load_data()
        clear_input()
    except mysql.connector.IntegrityError:
        messagebox.showerror("Lỗi", f"Mã giáo viên '{ma}' đã tồn tại!")
    except Exception as e:
        messagebox.showerror("Lỗi", str(e))
    finally:
        conn.close()

def XoaGiaoVien():
    selected = tree.selection()
    if not selected:
        messagebox.showwarning("Chưa chọn", "Hãy chọn giáo viên trên bảng để xóa")
        return
    ma = tree.item(selected)["values"][0]

    confirm = messagebox.askyesno("Xác nhận", f"Bạn có chấc muốn xóa giáo viên có mã {ma}?")
    if confirm:
        conn = connect_db()
        cur = conn.cursor()
        cur.execute("DELETE FROM giaovien WHERE ma_gv = %s", (ma,))
        conn.commit()
        conn.close()
        load_data()
        clear_input()
        messagebox.showinfo("Thành công", "Đã xóa giáo viên")

def SuaGiaoVien(event=None):
    selected = tree.selection()
    if not selected:
        if event is None:
            messagebox.showwarning("Chưa chọn", "Hãy chọn giáo viên để sửa")
        return
    
    values = tree.item(selected)["values"]

    entry_ma.delete(0, tk.END)
    entry_ma.insert(0, values[0])
    entry_ma.config(state='readonly')

    entry_holot.delete(0, tk.END)
    entry_holot.insert(0, values[1])

    entry_ten.delete(0, tk.END)
    entry_ten.insert(0, values[2])

    gender_var.set(values[3])
    date_entry.set_date(values[4])
    cbb_mon.set(values[5])

def LuuGiaoVien():
    ma = entry_ma.get()
    holot = entry_holot.get()
    ten = entry_ten.get()
    gioitinh = gender_var.get()
    ngaysinh = date_entry.get_date()
    monday = cbb_mon.get()

    if ma == "":
        messagebox.showwarning("Cảnh báo", "Vui lòng chọn Sửa một giáo viên trước khi Lưu")
        return
    conn = connect_db()
    try:
        cur = conn.cursor()
        sql = """UPDATE giaovien
                 SET ho_lot = %s, ten = %s, gioitinh = %s, ngay_sinh = %s, mon_day = %s
                 WHERE ma_gv = %s"""
        val = (holot, ten, gioitinh, ngaysinh, monday, ma)
        cur.execute(sql, val)
        conn.commit()
        messagebox.showinfo("Thành công", "Cập nhật thông tin giáo viên thành công")
        load_data()
        clear_input()
    except Exception as e:
        messagebox.showerror("Lỗi", str(e))
    finally:
        conn.close()

def TimKiem():
    search_win = tk.Toplevel(root)
    search_win.title("Tìm kiếm giáo viên")
    search_win.geometry("300x120")

    tk.Label(search_win, text="Nhập mã hoặc tên giáo viên:").pack(pady=10)
    entry_search = tk.Entry(search_win, width=30)
    entry_search.pack(pady=5)
    def ThucHienTim():
        keyword = entry_search.get()
        if keyword == "":
            messagebox.showwarning("Thông báo", "Vui lòng nhập từ khóa!")
            return
        for i in tree.get_children():
            tree.delete(i)
        conn = connect_db()
        try:
            cur = conn.cursor()
            sql = "SELECT * FROM giaovien WHERE ma_gv LIKE %s OR ten LIKE %s OR ho_lot LIKE %s"
            val = (f"%{keyword}%", f"%{keyword}%", f"%{keyword}%")
            cur.execute(sql, val)
            rows = cur.fetchall()
            if len(rows) == 0:
                messagebox.showinfo("Kết quả", "Không tìm thấy giáo viên nào.")
                load_data()
            else:
                for row in rows:
                    tree.insert("", tk.END, values=row)
            search_win.destroy()
        except Exception as e:
            messagebox.showerror("Lỗi", str(e))
        finally:
            conn.close()

    tk.Button(search_win, text="Tìm kiếm", command=ThucHienTim).pack(pady=10)

def XuatExcel():
    conn = connect_db()
    try:
        # 1. Đổi đuôi file thành .xlsx
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                                                 title="Lưu file Excel")
        if file_path:
            cur = conn.cursor()
            cur.execute("SELECT * FROM giaovien")
            rows = cur.fetchall()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Danh Sách Giáo Viên"
            
            headers = ["Mã GV", "Họ Lót", "Tên", "Giới Tính", "Ngày Sinh", "Môn Dạy"]
            ws.append(headers)
            
            header_font = Font(bold=True, color="FFFFFF")
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

                cell.fill = openpyxl.styles.PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")

            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))
            
            for row in rows:
                ws.append(row)
                for cell in ws[ws.max_row]:
                    cell.border = thin_border
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length + 5
            wb.save(file_path)
            messagebox.showinfo("Thành công", f"Đã xuất file Excel tại:\n{file_path}")
            
    except Exception as e:
        messagebox.showerror("Lỗi", str(e))
    finally:
        conn.close()
        
root = tk.Tk()
root.title("Quản lý giáo viên phổ thông")
center_window(root, 750, 550)
root.resizable(False, False)

lbl_title = tk.Label(root, text="Quản Lý Giáo Viên Phổ Thông", font=("Arial", 20, "bold"), fg="#d35400")
lbl_title.pack(pady=10)

frame_info = tk.Frame(root)
frame_info.pack(pady=5, padx=10, fill="x")

tk.Label(frame_info, text="Mã GV:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
entry_ma = tk.Entry(frame_info, width=15)
entry_ma.grid(row=0, column=1, padx=5, pady=5, sticky="w")

tk.Label(frame_info, text="Môn dạy:").grid(row=0, column=2, padx=5, pady=5, sticky="w")

cbb_mon = ttk.Combobox(frame_info, values=["Toán", "Vật Lý", "Hóa Học", "Sinh Học", "Ngữ Văn", "Tiếng Anh", "Lịch Sử", "Địa Lý", "Tin Học", "GDCD"], width=18, state="readonly")
cbb_mon.grid(row=0, column=3, padx=5, pady=5, sticky="w")

tk.Label(frame_info, text="Họ lót:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
entry_holot = tk.Entry(frame_info, width=25)
entry_holot.grid(row=1, column=1, padx=5, pady=5, sticky="w")

tk.Label(frame_info, text="Tên:").grid(row=1, column=2, padx=5, pady=5, sticky="w")
entry_ten = tk.Entry(frame_info, width=15)
entry_ten.grid(row=1, column=3, padx=5, pady=5, sticky="w")

tk.Label(frame_info, text="Giới tính:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
gender_var = tk.StringVar(value="Nam")
frame_gender = tk.Frame(frame_info)
frame_gender.grid(row=2, column=1, sticky="w")
tk.Radiobutton(frame_gender, text="Nam", variable=gender_var, value="Nam").pack(side=tk.LEFT)
tk.Radiobutton(frame_gender, text="Nữ", variable=gender_var, value="Nữ").pack(side=tk.LEFT, padx=10)

tk.Label(frame_info, text="Ngày sinh:").grid(row=2, column=2, padx=5, pady=5, sticky="w")

date_entry = DateEntry(frame_info, width=16, background="darkblue", 
                       foreground="white", date_pattern="yyyy-mm-dd",
                       year=2000, month=1, day=1)
date_entry.grid(row=2, column=3, padx=5, pady=5, sticky="w")

frame_btn = tk.Frame(root)
frame_btn.pack(pady=15)

btn_width = 9
tk.Button(frame_btn, text="Thêm", width=btn_width, command=ThemGiaoVien, bg="#2ecc71", fg="white").grid(row=0, column=0, padx=5)
tk.Button(frame_btn, text="Lưu", width=btn_width, command=LuuGiaoVien, bg="#3498db", fg="white").grid(row=0, column=1, padx=5)
tk.Button(frame_btn, text="Sửa", width=btn_width, command=SuaGiaoVien, bg="#f1c40f").grid(row=0, column=2, padx=5)
tk.Button(frame_btn, text="Hủy", width=btn_width, command=clear_input).grid(row=0, column=3, padx=5)
tk.Button(frame_btn, text="Xóa", width=btn_width, command=XoaGiaoVien, bg="#e74c3c", fg="white").grid(row=0, column=4, padx=5)
tk.Button(frame_btn, text="Tìm Kiếm", width=btn_width, command=TimKiem, bg="#9b59b6", fg="white").grid(row=0, column=6, padx=5)
tk.Button(frame_btn, text="Xem Tất Cả", width=btn_width, command=load_data, bg="#95a5a6", fg="white").grid(row=0, column=7, padx=5)
tk.Button(frame_btn, text="Xuất Excel", width=btn_width, command=XuatExcel, bg="#1abc9c", fg="white").grid(row=0, column=8, padx=5)
tk.Button(frame_btn, text="Thoát", width=btn_width, command=root.quit).grid(row=0, column=9, padx=5)

tk.Label(root, text="Danh sách giáo viên", font=("Arial", 10, "bold")).pack(pady=5, anchor="w", padx=20)

columns = ("ma_gv", "holot", "ten", "gioitinh", "ngaysinh", "monday")
tree = ttk.Treeview(root, columns=columns, show="headings", height=10)

tree.heading("ma_gv", text="Mã GV")
tree.heading("holot", text="Họ lót")
tree.heading("ten", text="Tên")
tree.heading("gioitinh", text="Giới tính")
tree.heading("ngaysinh", text="Ngày sinh")
tree.heading("monday", text="Môn dạy")

tree.column("ma_gv", width=80, anchor="center")
tree.column("holot", width=150)
tree.column("ten", width=80)
tree.column("gioitinh", width=60, anchor="center")
tree.column("ngaysinh", width=100, anchor="center")
tree.column("monday", width=120)

tree.pack(padx=20, pady=5, fill="both", expand=True)

tree.bind("<Double-1>", SuaGiaoVien)

load_data()
clear_input()
root.mainloop()